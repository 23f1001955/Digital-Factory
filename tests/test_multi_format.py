import json
import os
import shutil

from agents.csv_export_agent import run
from agents import market_agent
from orchestrator.models import ComponentSpec, JobSpec


def _clean_output(slug: str):
    path = os.path.join("outputs", slug)
    if os.path.exists(path):
        shutil.rmtree(path)


def test_csv_export_agent_multi_format(tmp_path):
    """csv_export_agent writes both CSV and XLSX when active_formats has both."""
    research_data = {
        "niche": "real estate agents",
        "database": [
            {"name": "Agent A", "city": "NYC", "phone": "555-0101"},
            {"name": "Agent B", "city": "LA", "phone": "555-0102"},
        ]
    }
    research_path = tmp_path / "research.json"
    with open(research_path, "w") as f:
        json.dump(research_data, f)

    comp = ComponentSpec(
        id="database_export",
        agent="csv_export_agent",
        output="data/test_multi_db.csv",
        depends_on=["market_research"],
        delivery=["zip"],
        capabilities=["csv", "xlsx"],
        active_formats=["csv", "xlsx"],
    )
    job = JobSpec(slug="test-multi", product_type="database", niche="real estate")
    context = {"market_research": str(research_path)}

    _clean_output("test-multi")
    result = run(comp, job, context)
    assert result.status == "done"
    assert result.output_paths is not None
    assert "csv" in result.output_paths
    assert "xlsx" in result.output_paths

    with open(result.output_paths["csv"]) as f:
        content = f.read()
    assert "Agent A" in content
    assert "NYC" in content

    import openpyxl
    wb = openpyxl.load_workbook(result.output_paths["xlsx"])
    ws = wb.active
    assert ws["A1"].value == "name"
    assert ws["A2"].value == "Agent A"
    assert ws["B2"].value == "NYC"


def test_market_agent_recommends_formats(monkeypatch):
    """Market_agent includes recommended_formats when LLM provides them."""
    monkeypatch.setattr(
        "agents.llm_client.generate_text",
        lambda p: json.dumps({
            "competitor_landscape": {"direct_competitors": [], "pricing_tiers": {}, "recommended_price": 29, "quality_gaps": [], "trending_keywords": []},
            "content_recommendations": {"tone": "professional", "key_themes": [], "seo_keywords": []},
            "market_insights": {},
            "recommended_product_type": "database",
            "recommendation_confidence": 0.85,
            "recommendation_reasoning": "High demand for lead lists",
            "recommended_formats": {
                "database_export": ["csv", "xlsx"],
                "market_research": ["pdf"],
            },
            "pipeline_plan": {"components": []},
        }),
    )

    monkeypatch.setattr("agents.research_tools.brave_search", lambda q, n: [])
    monkeypatch.setattr("agents.research_tools.duckduckgo_search", lambda q, n: [])
    monkeypatch.setattr("agents.research_tools.reddit_search", lambda q, n: [])
    monkeypatch.setattr("agents.research_tools.gdelt_news", lambda q, n: [])
    monkeypatch.setattr("agents.research_tools.newsapi_headlines", lambda q, n: [])
    monkeypatch.setattr("agents.research_tools.pytrends_data", lambda q: {})
    monkeypatch.setattr("agents.research_tools.firecrawl_scrape", lambda u: None)

    job_spec = JobSpec(slug="test-format-rec", product_type="database", niche="real estate")
    comp = ComponentSpec(id="market_research", agent="market_agent", output="data/market_research.json", depends_on=[])
    context = {}

    _clean_output("test-format-rec")
    result = market_agent.run(comp, job_spec, context)
    assert result.status == "done"

    with open(result.output_path) as f:
        research = json.load(f)

    assert "recommended_formats" in research
    assert "database_export" in research["recommended_formats"]
    assert "xlsx" in research["recommended_formats"]["database_export"]


def test_market_agent_fallback_formats(monkeypatch):
    """Market_agent fallback includes empty recommended_formats."""
    monkeypatch.setattr(
        "agents.llm_client.generate_text",
        lambda p: (_ for _ in ()).throw(Exception("LLM unavailable")),
    )

    monkeypatch.setattr("agents.research_tools.brave_search", lambda q, n: [])
    monkeypatch.setattr("agents.research_tools.duckduckgo_search", lambda q, n: [])
    monkeypatch.setattr("agents.research_tools.reddit_search", lambda q, n: [])
    monkeypatch.setattr("agents.research_tools.gdelt_news", lambda q, n: [])
    monkeypatch.setattr("agents.research_tools.newsapi_headlines", lambda q, n: [])
    monkeypatch.setattr("agents.research_tools.pytrends_data", lambda q: {})
    monkeypatch.setattr("agents.research_tools.firecrawl_scrape", lambda u: None)

    job_spec = JobSpec(slug="test-fallback-rec", product_type="discovery", niche="test")
    comp = ComponentSpec(id="market_research", agent="market_agent", output="data/market_research.json", depends_on=[])
    context = {}

    _clean_output("test-fallback-rec")
    result = market_agent.run(comp, job_spec, context)
    assert result.status == "done"

    with open(result.output_path) as f:
        research = json.load(f)

    assert "recommended_formats" in research
    assert research["recommended_formats"] == {}


def test_orchestrator_merges_format_recs(tmp_path, monkeypatch):
    """Orchestrator sets active_formats after market_agent recommends."""
    from orchestrator.orchestrator import Orchestrator
    from orchestrator.state import load_job_state
    from orchestrator.models import AgentResult, ProductSchema
    from agents.registry import AGENT_REGISTRY
    from unittest import mock
    import json
    import os
    import shutil

    slug = "test-merge-fmts"

    def _write_json(path, data):
        path = tmp_path / path
        path.parent.mkdir(parents=True, exist_ok=True)
        with open(path, "w") as f:
            json.dump(data, f)

    _write_json("job_spec.json", {
        "slug": slug, "product_type": "database", "niche": "test",
        "notion_sync": False, "notion_parent_page_id": None,
        "created_at": "2026-06-16T10:00:00Z",
    })

    _write_json("database.json", {
        "product_type": "database", "display_name": "Database",
        "components": [
            {"id": "market_research", "agent": "market_agent", "output": "data/market_research.json", "depends_on": []},
            {"id": "database_export", "agent": "csv_export_agent", "output": "data/{slug}_database.csv", "depends_on": ["market_research"], "delivery": ["zip"], "capabilities": ["csv", "xlsx"]},
        ],
    })

    def mock_market(comp, js, ctx):
        out_dir = os.path.join("outputs", js.slug, "data")
        os.makedirs(out_dir, exist_ok=True)
        research = {
            "niche": "test",
            "recommended_product_type": "database",
            "recommendation_confidence": 0.9,
            "recommendation_reasoning": "test",
            "recommended_formats": {"database_export": ["csv", "xlsx"]},
            "pipeline_plan": {"components": []},
        }
        path = os.path.join(out_dir, "market_research.json")
        with open(path, "w") as f:
            json.dump(research, f)
        return AgentResult(status="done", output_path=path)

    monkeypatch.setitem(AGENT_REGISTRY, "market_agent", mock_market)
    monkeypatch.setitem(AGENT_REGISTRY, "csv_export_agent", mock.Mock(return_value=AgentResult(status="done")))

    out_path = os.path.join("outputs", slug)
    if os.path.exists(out_path):
        shutil.rmtree(out_path)

    orc = Orchestrator(str(tmp_path / "job_spec.json"))
    with open(tmp_path / "database.json") as f:
        orc.schema = ProductSchema(**json.load(f))
    orc.state_path = os.path.join("outputs", slug, "job_state.json")
    orc.state = load_job_state(orc.state_path, slug)
    monkeypatch.setattr(orc, "_generate_run_summary", lambda: None)

    orc.run()

    export_comp = [c for c in orc.schema.components if c.id == "database_export"]
    assert len(export_comp) == 1
    assert "csv" in export_comp[0].active_formats
    assert "xlsx" in export_comp[0].active_formats


def test_invalid_format_filtered(tmp_path, monkeypatch):
    """Unknown formats in recommended_formats are silently filtered."""
    from orchestrator.orchestrator import Orchestrator
    from orchestrator.state import load_job_state
    from orchestrator.models import AgentResult, ProductSchema
    from agents.registry import AGENT_REGISTRY
    from unittest import mock
    import json
    import os
    import shutil

    slug = "test-filter-fmts"

    def _write_json(path, data):
        path = tmp_path / path
        path.parent.mkdir(parents=True, exist_ok=True)
        with open(path, "w") as f:
            json.dump(data, f)

    _write_json("job_spec.json", {
        "slug": slug, "product_type": "database", "niche": "test",
        "notion_sync": False, "notion_parent_page_id": None,
        "created_at": "2026-06-16T10:00:00Z",
    })

    _write_json("database.json", {
        "product_type": "database", "display_name": "Database",
        "components": [
            {"id": "market_research", "agent": "market_agent", "output": "data/market_research.json", "depends_on": []},
            {"id": "database_export", "agent": "csv_export_agent", "output": "data/{slug}_database.csv", "depends_on": ["market_research"], "delivery": ["zip"], "capabilities": ["csv"]},
        ],
    })

    def mock_market(comp, js, ctx):
        out_dir = os.path.join("outputs", js.slug, "data")
        os.makedirs(out_dir, exist_ok=True)
        research = {
            "niche": "test",
            "recommended_product_type": "database",
            "recommendation_confidence": 0.9,
            "recommendation_reasoning": "test",
            "recommended_formats": {"database_export": ["csv", "xlsx", "pdf"]},
            "pipeline_plan": {"components": []},
        }
        path = os.path.join(out_dir, "market_research.json")
        with open(path, "w") as f:
            json.dump(research, f)
        return AgentResult(status="done", output_path=path)

    monkeypatch.setitem(AGENT_REGISTRY, "market_agent", mock_market)
    monkeypatch.setitem(AGENT_REGISTRY, "csv_export_agent", mock.Mock(return_value=AgentResult(status="done")))

    out_path = os.path.join("outputs", slug)
    if os.path.exists(out_path):
        shutil.rmtree(out_path)

    orc = Orchestrator(str(tmp_path / "job_spec.json"))
    with open(tmp_path / "database.json") as f:
        orc.schema = ProductSchema(**json.load(f))
    orc.state_path = os.path.join("outputs", slug, "job_state.json")
    orc.state = load_job_state(orc.state_path, slug)
    monkeypatch.setattr(orc, "_generate_run_summary", lambda: None)

    orc.run()

    export_comp = [c for c in orc.schema.components if c.id == "database_export"]
    assert len(export_comp) == 1
    assert export_comp[0].active_formats == ["csv"]


def test_no_recs_legacy_mode(tmp_path, monkeypatch):
    """No recommended_formats means legacy mode — no active_formats set."""
    from orchestrator.orchestrator import Orchestrator
    from orchestrator.state import load_job_state
    from orchestrator.models import AgentResult, ProductSchema
    from agents.registry import AGENT_REGISTRY
    from unittest import mock
    import json
    import os
    import shutil

    slug = "test-legacy-fmts"

    def _write_json(path, data):
        path = tmp_path / path
        path.parent.mkdir(parents=True, exist_ok=True)
        with open(path, "w") as f:
            json.dump(data, f)

    _write_json("job_spec.json", {
        "slug": slug, "product_type": "database", "niche": "test",
        "notion_sync": False, "notion_parent_page_id": None,
        "created_at": "2026-06-16T10:00:00Z",
    })

    _write_json("database.json", {
        "product_type": "database", "display_name": "Database",
        "components": [
            {"id": "market_research", "agent": "market_agent", "output": "data/market_research.json", "depends_on": []},
            {"id": "database_export", "agent": "csv_export_agent", "output": "data/{slug}_database.csv", "depends_on": ["market_research"], "delivery": ["zip"], "capabilities": ["csv", "xlsx"]},
        ],
    })

    def mock_market(comp, js, ctx):
        out_dir = os.path.join("outputs", js.slug, "data")
        os.makedirs(out_dir, exist_ok=True)
        research = {
            "niche": "test",
            "recommended_product_type": "database",
            "recommendation_confidence": 0.9,
            "recommendation_reasoning": "test",
            "pipeline_plan": {"components": []},
        }
        path = os.path.join(out_dir, "market_research.json")
        with open(path, "w") as f:
            json.dump(research, f)
        return AgentResult(status="done", output_path=path)

    monkeypatch.setitem(AGENT_REGISTRY, "market_agent", mock_market)
    monkeypatch.setitem(AGENT_REGISTRY, "csv_export_agent", mock.Mock(return_value=AgentResult(status="done")))

    out_path = os.path.join("outputs", slug)
    if os.path.exists(out_path):
        shutil.rmtree(out_path)

    orc = Orchestrator(str(tmp_path / "job_spec.json"))
    with open(tmp_path / "database.json") as f:
        orc.schema = ProductSchema(**json.load(f))
    orc.state_path = os.path.join("outputs", slug, "job_state.json")
    orc.state = load_job_state(orc.state_path, slug)
    monkeypatch.setattr(orc, "_generate_run_summary", lambda: None)

    orc.run()

    export_comp = [c for c in orc.schema.components if c.id == "database_export"]
    assert len(export_comp) == 1
    assert export_comp[0].active_formats == []


def test_delivery_map_per_format(tmp_path, monkeypatch):
    """delivery_map stores per-format paths from agent_result.output_paths."""
    from orchestrator.orchestrator import Orchestrator
    from orchestrator.state import load_job_state
    from orchestrator.models import AgentResult, ProductSchema
    from agents.registry import AGENT_REGISTRY
    import json
    import os
    import shutil

    slug = "test-delivery-fmts"

    def _write_json(path, data):
        path = tmp_path / path
        path.parent.mkdir(parents=True, exist_ok=True)
        with open(path, "w") as f:
            json.dump(data, f)

    _write_json("job_spec.json", {
        "slug": slug, "product_type": "database", "niche": "test",
        "notion_sync": False, "notion_parent_page_id": None,
        "created_at": "2026-06-16T10:00:00Z",
    })

    _write_json("database.json", {
        "product_type": "database", "display_name": "Database",
        "components": [
            {"id": "market_research", "agent": "market_agent", "output": "data/market_research.json", "depends_on": []},
            {"id": "database_export", "agent": "csv_export_agent", "output": "data/{slug}_database.csv", "depends_on": ["market_research"], "delivery": ["zip"], "capabilities": ["csv", "xlsx"]},
        ],
    })

    def mock_market(comp, js, ctx):
        out_dir = os.path.join("outputs", js.slug, "data")
        os.makedirs(out_dir, exist_ok=True)
        path = os.path.join(out_dir, "market_research.json")
        with open(path, "w") as f:
            json.dump({"niche": "test", "pipeline_plan": {"components": []}}, f)
        return AgentResult(status="done", output_path=path)

    def mock_csv(comp, js, ctx):
        out_dir = os.path.join("outputs", js.slug, "data")
        os.makedirs(out_dir, exist_ok=True)
        csv_path = os.path.join(out_dir, "database_export.csv")
        xlsx_path = os.path.join(out_dir, "database_export.xlsx")
        with open(csv_path, "w") as f:
            f.write("name,city\nA,NYC\n")
        return AgentResult(
            status="done",
            output_path=csv_path,
            output_paths={"csv": csv_path, "xlsx": xlsx_path},
        )

    monkeypatch.setitem(AGENT_REGISTRY, "market_agent", mock_market)
    monkeypatch.setitem(AGENT_REGISTRY, "csv_export_agent", mock_csv)

    out_path = os.path.join("outputs", slug)
    if os.path.exists(out_path):
        shutil.rmtree(out_path)

    orc = Orchestrator(str(tmp_path / "job_spec.json"))
    with open(tmp_path / "database.json") as f:
        orc.schema = ProductSchema(**json.load(f))
    orc.state_path = os.path.join("outputs", slug, "job_state.json")
    orc.state = load_job_state(orc.state_path, slug)
    monkeypatch.setattr(orc, "_generate_run_summary", lambda: None)

    orc.run()

    dm = orc._build_delivery_map()
    assert "database_export" in dm
    outputs = dm["database_export"].get("outputs", {})
    assert "csv" in outputs
    assert "xlsx" in outputs
    assert outputs["csv"].endswith(".csv")
    assert outputs["xlsx"].endswith(".xlsx")
